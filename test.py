import os
import numpy as np
import pandas as pd
import torchvision
from torch.nn import Softmax
from model import AudioDFNet
import torch
from PIL import Image
import openpyxl

workbook = openpyxl.Workbook()
sheet = workbook.create_sheet('Sheet')

eval_path = 'D:/DataSet/ASVspoof2019/eval_fig/'
fig_list = os.listdir(eval_path)
eval_label = pd.read_csv('D:/DataSet/ASVspoof2019/ASVspoof2019.LA.cm.eval.trl.csv', header=None)

model = torch.load('./model/AudioDFNet46.pt', map_location=torch.device('cpu'))

softmax = Softmax(1)

transform = torchvision.transforms.Compose([torchvision.transforms.ToTensor(),
                                            torchvision.transforms.Resize((32, 32))])


def labelConvert(spoof_type):
    if spoof_type == 'A07' or spoof_type == 'A08' or spoof_type == 'A09' or spoof_type == 'A10' or spoof_type == 'A11' \
            or spoof_type == 'A12' or spoof_type == 'A16':
        return 1
    elif spoof_type == 'A17' or spoof_type == 'A18' or spoof_type == 'A19' or spoof_type == 'A13' \
            or spoof_type == 'A14' or spoof_type == 'A15':
        return 2
    else:
        return 0

count1 = 0
count2 = 0

model.eval()
with torch.no_grad():
    for i in range(len(fig_list)):
        real_type = labelConvert(eval_label.iloc[i, 3])
        if real_type == 0:
            count1 += 1
        else:
            continue
        # eval
        image = Image.open(eval_path + fig_list[i])
        image = transform(image)
        image = image.reshape(1, 3, 32, 32)
        output = model(image)
        confidence = softmax(output)
        confidence = confidence.numpy().flatten()
        eval_type = np.argmax(confidence)
        # compare
        if eval_type == 0:
            count2 += 1
        print('eval type: {}, real type: {}'.format(eval_type, real_type))
        # sheet.cell(i + 1, 1, str(eval_type))
        # sheet.cell(i + 1, 2, str(real_type))

# workbook.save('eval_result2.xlsx')
print(count2 / count1)